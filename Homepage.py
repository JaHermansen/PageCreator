import streamlit as st
import pandas as pd
import io
from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string
import os
import shutil
from PIL import Image

downloads_path = os.path.join(os.path.expanduser("~"), "Downloads")
# Page icon
icon = Image.open('paa1.png')


def load_excel(file, sheet):
    return pd.read_excel(file, sheet_name=sheet)


def remove_uploaded_files():
    if "file_name" in st.session_state:
        del st.session_state["file_name"]
    if "is_file_uploaded" in st.session_state:
        del st.session_state["is_file_uploaded"]
    if "template_file_name" in st.session_state:
        del st.session_state["template_file_name"]
    if "template_file_uploaded" in st.session_state:
        del st.session_state["template_file_uploaded"]
    if "df_extracted_rows" in st.session_state:
        del st.session_state["df_extracted_rows"]


def display_upload(uploaded_file):
    if uploaded_file:
        xls = pd.ExcelFile(uploaded_file)
        sheet_names = xls.sheet_names
        sheet = st.sidebar.selectbox("Choose a sheet", sheet_names)
        df = load_excel(uploaded_file, sheet)
        st.session_state["file_name"] = uploaded_file.name
        st.session_state["is_file_uploaded"] = True

        start_row = st.sidebar.number_input('Start row', min_value=1, value=1) - 2  # adjusting for 0 index
        end_row = st.sidebar.number_input('End row', min_value=start_row + 1, value=start_row + 1) - 1  # adjusting for 0 index
        st.session_state["df_extracted_rows"] = df.iloc[start_row:end_row]

    if st.sidebar.button("Preview selected rows"):
        st.write(st.session_state["df_extracted_rows"])


def generate_files(df_extracted_rows, template_file_path, output_path=downloads_path, text_input1="", text_input2="", text_input3=""):
    os.makedirs(output_path, exist_ok=True)

    cell_mapping = {
        'E': 'D37',
        'F': 'D38',
        'V': 'D39',
        'U': ['D40', 'B45'],
        'W': 'C45',
        'AC': 'H45',
        'AD': 'I45',
    }

    # Add a new function to overwrite cells based on user inputs
    def overwrite_cells(sheet, cell, value):
        if isinstance(cell, list):
            for c in cell:
                overwrite_cells(sheet, c, value)
        else:
            merged_ranges = sheet.merged_cells.ranges
            for merged_range in merged_ranges:
                if cell in merged_range:
                    sheet.unmerge_cells(str(merged_range))
                    min_col, min_row, max_col, max_row = merged_range.bounds
                    for row in sheet.iter_rows(min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col):
                        for cell_in_range in row:
                            cell_in_range.value = value
                    sheet.merge_cells(start_row=min_row, end_row=max_row, start_column=min_col, end_column=max_col)
                    break
            else:
                sheet[cell] = value

    for i, row in df_extracted_rows.iterrows():
        file_name = row[column_index_from_string('F') - 1]  # Get the value in cell F
        new_file = os.path.join(output_path, f'{file_name}.xlsx')  # Use the value in cell F as the file name
        shutil.copyfile(template_file_path, new_file)

        book = load_workbook(new_file)
        sheet = book.active

        for df_col, excel_cells in cell_mapping.items():
            df_col_idx = column_index_from_string(df_col) - 1  # Convert to 0-based index
            cell_value = row[df_col_idx]

            if not isinstance(excel_cells, list):
                excel_cells = [excel_cells]

            for excel_cell in excel_cells:
                overwrite_cells(sheet, excel_cell, cell_value)

        # Overwrite cells based on user inputs
        overwrite_cells(sheet, 'D20', text_input1)
        overwrite_cells(sheet, 'D22', text_input2)
        overwrite_cells(sheet, 'D23', text_input3)

        book.save(new_file)


def main():
    st.set_page_config(
        layout="wide",
        page_title="PAA Excel Page Creator",
        page_icon=icon,
    )
    st.markdown("<h1 style='color: #006095;'>Excel Front Page Generator</h1>", unsafe_allow_html=True)
    st.markdown("### Click on Browse File in the Side Bar to start")

    uploaded_file = st.sidebar.file_uploader("Choose a file", type=["xlsx", "xls"], key="uploaded_file")

    display_upload(uploaded_file)

    template_file = st.sidebar.file_uploader("Upload the Page Template", type=["xlsx", "xls"], key="template_file")

    if template_file:
        st.session_state["template_file_name"] = template_file.name
        template_file_path = os.path.join(os.getcwd(), template_file.name)  # Set the template file path
        st.session_state["template_file_path"] = template_file_path
        with open(template_file_path, "wb") as f:
            f.write(template_file.getvalue())

    if uploaded_file is None and "uploaded_file" in st.session_state and template_file is None and "template_file" in st.session_state:
        remove_uploaded_files()

    if uploaded_file and template_file:
        #st.sidebar.success("Project successfully loaded")

        st.subheader("Enter Values for Overwriting Cells")
        text_input1 = st.text_input("Slut dokumentation")  # to cell B-I 20
        text_input2 = st.text_input("Sektion")  # to cell B-I 22
        text_input3 = st.text_input("Delområde")  # to cell B-I 23

        if st.button('Generate Files'):
            generate_files(
                st.session_state["df_extracted_rows"],
                template_file_path=st.session_state["template_file_path"],
                text_input1=text_input1,
                text_input2=text_input2,
                text_input3=text_input3
            )
    else:
        if not uploaded_file:
            st.warning("Please upload a data file first.")
        elif not template_file:
            st.warning("Please upload a template file to generate pages.")


if __name__ == "__main__":
    main()
